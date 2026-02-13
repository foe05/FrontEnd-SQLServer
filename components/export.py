"""
Excel export functionality for time tracking data
"""
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime
from typing import Dict, Any

class ExcelExporter:
    """Excel export manager with formatting"""
    
    def __init__(self):
        self.status_colors = {
            '🟢 Buchbar': '92D050',      # Green
            '🟡 Kritisch': 'FFFF00',     # Yellow  
            '🔴 Überbucht': 'FF0000'     # Red
        }
    
    def create_workbook(self, data: pd.DataFrame, metadata: Dict[str, Any] = None) -> BytesIO:
        """Create formatted Excel workbook"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Zeiterfassung Dashboard"
        
        # Add metadata header
        if metadata:
            self._add_metadata_header(ws, metadata)
            start_row = 8
        else:
            start_row = 1
        
        # Add data
        self._add_data_table(ws, data, start_row)
        
        # Format worksheet
        self._format_worksheet(ws, data, start_row)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _add_metadata_header(self, ws, metadata: Dict[str, Any]):
        """Add metadata header to worksheet"""
        # Title
        ws['A1'] = "SQL Server Dashboard - Zeiterfassung Export"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Export info
        ws['A3'] = f"Export erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A4'] = f"Benutzer: {metadata.get('user', 'Unbekannt')}"
        # Filter None values before join
        projects = [str(p) for p in metadata.get('projects', []) if p is not None]
        ws['A5'] = f"Projekte: {', '.join(projects) if projects else 'Keine'}"
        ws['A6'] = f"Zeitraum: {metadata.get('period', 'Alle')}"
        
        # Styling
        for cell in [ws['A3'], ws['A4'], ws['A5'], ws['A6']]:
            cell.font = Font(size=10)
    
    def _add_data_table(self, ws, data: pd.DataFrame, start_row: int):
        """Add data table to worksheet"""
        if data.empty:
            ws[f'A{start_row}'] = "Keine Daten verfügbar"
            return
        
        # Add headers
        headers = data.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Add data rows
        for row_idx, row_data in enumerate(data.itertuples(index=False), start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply status coloring if Status column
                if col_idx <= len(headers) and headers[col_idx-1] == 'Status':
                    self._apply_status_formatting(cell, str(value))
    
    def _apply_status_formatting(self, cell, status: str):
        """Apply color formatting based on status"""
        color = self.status_colors.get(status, 'FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # White text for red background
        if color == 'FF0000':
            cell.font = Font(color='FFFFFF', bold=True)
        elif color in ['FFFF00', '92D050']:
            cell.font = Font(bold=True)
    
    def _format_worksheet(self, ws, data: pd.DataFrame, start_row: int):
        """Format the entire worksheet"""
        if data.empty:
            return
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(data.columns, 1):
            max_length = len(str(column))
            
            # Check data column for max length
            for value in data.iloc[:, col_idx-1]:
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            
            # Set column width (max 50 characters)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # Add borders to data table
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(data), 
                              min_col=1, max_col=len(data.columns)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
    
    def export_summary_data(self, df: pd.DataFrame, metadata: Dict[str, Any] = None) -> BytesIO:
        """Export aggregated summary data"""
        if df.empty:
            return self.create_workbook(pd.DataFrame({"Info": ["Keine Daten verfügbar"]}), metadata)
        
        # Create summary DataFrame
        summary_data = df.copy()
        
        # Round numeric columns
        numeric_cols = ['ActualHours', 'TargetHours', 'Fulfillment']
        for col in numeric_cols:
            if col in summary_data.columns:
                summary_data[col] = pd.to_numeric(summary_data[col], errors='coerce').round(2)
        
        # Rename columns to German
        column_mapping = {
            'Activity': 'Tätigkeit',
            'TargetHours': 'Sollstunden',
            'ActualHours': 'Iststunden', 
            'Fulfillment': 'Erfüllungsstand (%)',
            'Status': 'Status',
            'Projekt': 'Projekt',
            'Kundenname': 'Kunde'
        }
        
        summary_data = summary_data.rename(columns=column_mapping)
        
        return self.create_workbook(summary_data, metadata)
    
    def export_detailed_data(self, df: pd.DataFrame, metadata: Dict[str, Any] = None) -> BytesIO:
        """Export detailed time entries"""
        if df.empty:
            return self.create_workbook(pd.DataFrame({"Info": ["Keine Daten verfügbar"]}), metadata)
        
        # Select and rename relevant columns
        detail_columns = {
            'Name': 'Mitarbeiter',
            'Zeit': 'Stunden',
            'Projekt': 'Projekt', 
            'Verwendung': 'Tätigkeit',
            'Status': 'Status',
            'Datum': 'Datum',
            'Kommentar': 'Kommentar',
            'Kundenname': 'Kunde'
        }
        
        # Filter existing columns
        available_cols = {k: v for k, v in detail_columns.items() if k in df.columns}
        detail_data = df[list(available_cols.keys())].rename(columns=available_cols)
        
        return self.create_workbook(detail_data, metadata)
    
    def show_export_options(self, summary_df: pd.DataFrame, detail_df: pd.DataFrame = None, 
                          user_info: Dict[str, Any] = None, filters: Dict[str, Any] = None):
        """Show export options in Streamlit interface"""
        
        st.subheader("📥 Daten Export")
        
        col1, col2 = st.columns(2)
        
        # Prepare metadata
        metadata = {
            'user': user_info.get('name', 'Unbekannt') if user_info else 'Unbekannt',
            'projects': filters.get('selected_projects', []) if filters else [],
            'period': f"{filters.get('year', 'Alle')}" if filters else 'Alle'
        }
        
        with col1:
            # Summary export
            if st.button("📊 Zusammenfassung exportieren", type="primary"):
                if not summary_df.empty:
                    excel_buffer = self.export_summary_data(summary_df, metadata)
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                    filename = f"zeiterfassung_zusammenfassung_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="📥 Excel herunterladen",
                        data=excel_buffer.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success("Zusammenfassung Excel-Datei erstellt!")
                else:
                    st.warning("Keine Zusammenfassungsdaten zum Export verfügbar")
        
        with col2:
            # Detail export  
            if detail_df is not None and not detail_df.empty:
                if st.button("📋 Details exportieren", type="secondary"):
                    excel_buffer = self.export_detailed_data(detail_df, metadata)
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                    filename = f"zeiterfassung_details_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="📥 Excel herunterladen",
                        data=excel_buffer.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success("Detail Excel-Datei erstellt!")
            else:
                st.info("Detaildaten nicht verfügbar für Export")

# Global exporter instance
excel_exporter = ExcelExporter()
